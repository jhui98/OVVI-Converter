# Jacob Hui - Clover Automation Functions

from openpyxl import Workbook, load_workbook

def init_OVVI(folder):
    wb = Workbook() 
    dest_filename = folder + 'Item-PLU with Data.xlsx'   
    ws1 = wb.active
    ws1.title = "Item-PLU" 
    ws2 = wb.create_sheet(title="ModifierGroups")   
    
    
    wb.save(filename = dest_filename)
