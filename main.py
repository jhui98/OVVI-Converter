# Jacob Hui - Ovvi Converter

from methods import get_departments, item_department_dict, initialItemIstance, Ovvi, logo
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
from pathlib import Path
downloads_path = str(Path.home() / "Downloads")
print(downloads_path)

run = "y"
while run != "n":
    print(logo)

    task = input("""
    A: Clover Conversion
    B: Barcode Leading Zero Fix (Coming soon)
    C: Double Check My Work (Coming soon)

    Please choose an operation: """)

    if task == "A" or "a": # CLover to OVVI format change
        clover_path = input("\nPlease drag and drop the file, then press Enter: ")
        clover_path = clover_path[3:-1]
        # print(clover_path) # sanity check
        print("\nProcessing, please wait...\n")
        # import file path from clover
        # clover_folder = input(r"Input the folder path that you are using: ") 
        # clover_file = input(r"Input file name: ") + ".xlsx"
        # print(clover_folder, clover_file)

        # load worksheet
        clover_wb = load_workbook(clover_path)
        clover_ws = clover_wb.active
        # print(clover_wb.sheetnames) # sanity check

        departments = get_departments(clover_wb) # get departments

        items_dict = item_department_dict(clover_wb, departments) # create dictionary of items with assigned department
        items = initialItemIstance(clover_wb, departments)

        # double checking
        # for item in items:
        #     if item.itemName == '12CARLO ROSSI BURGUNDY 4L':
        #         print(item.itemName)
        #         print(item.itemDepartment)
        #         print(item.itemSellPrice)
        #         print(item.itemBarcode)
        #         print(item.itemCost)
        #         print(item.itemStock)
        
        # TODO: store in new sheet
        wb = Workbook()
        ws = wb.active
        # create Item-PLU sheet
        ws.title = "Item-PLU"
        ws.append(["Department", "ItemNumber", "ItemName", "ModifierGroups", "Description", "Barcode", "Cost", "SellPrice", "InStock", "Tax1", "DisplayInMenu", "IsInventoryItem", "IsFoodStampItem", "BeveragesDeposit"])

        # add in clover items
        for item in items:
            ws.append([item.itemDepartment, item.itemNumber, item.itemName, item.modifierGroups, item.description, item.itemBarcode, item.itemCost, item.itemSellPrice, item.inStock, item.tax1, item.displayInMenu, item.isInventoryItem, item.isFoodStampable, item.beverageDeposit])
        
        # create ModifierGroups sheet
        wb.create_sheet("ModifierGroups")
        ws = wb["ModifierGroups"]
        ws.append(["Modifer Group Department", "Modifier Group Name", "Charged", "Modifier Department", "Modifier", "Price", "Min", "Max"])

        timestr = time.strftime("%Y%m%d-%H%M%S")
        wb.save(downloads_path + f"\Ovvi_Convert_Output_{timestr}.xlsx")
        # TODO: save
    
    print("Finished! Please check your downloads folder for the updated file.\n")
    run = input("Would you like to process another operation? y or n: ")