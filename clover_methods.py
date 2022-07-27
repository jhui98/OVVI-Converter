# Jacob Hui - Clover Automation Functions

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from ovvi_object import Ovvi

def get_departments(clover_wb):
    ws = clover_wb["Categories"]
    departments = []
    for cell in ws[1]:
        departments.append(cell.value)
    departments.pop(0)
    # print(departments)
    return departments 

def item_department_dict(clover_wb, departments):
    item_dict = {}
    ws = clover_wb["Categories"] # select categories ws
    for col in range(len(departments)):
        dep = str(departments[col]) # save department name
        # print(dep)
        for cell in ws[get_column_letter(col+2)]: # for every cell within column index
            if (cell.value != None) and (cell.value != dep): # cell not empty and not department
                item = str(cell.value)
                item_dict[item] = {"department": dep}
    return item_dict

def initialItemIstance(clover_wb, departments):
    items = []
    ws = clover_wb["Categories"] # select categories ws
    for col in range(len(departments)):
        dep = str(departments[col]) # save department name
        for cell in ws[get_column_letter(col+2)]: # for every cell within column index
            if (cell.value != None) and (cell.value != dep): # cell not empty and not department
                item = str(cell.value)
                item = Ovvi(item, dep)
                items.append(item)

    ws = clover_wb["Items"] # select categories ws
    for row in ws.iter_rows(): # iter throguh rows
        rowIndex = str(row[0].row) # get a row number
        # get data from row based on col index
        Name = ws[f"B{rowIndex}"].value
        Price = ws[f"D{rowIndex}"].value
        Cost = ws[f"H{rowIndex}"].value
        Barcode = ws[f"J{rowIndex}"].value
        Stock = ws[f"L{rowIndex}"].value

        for item in items: # check items obj list for item name, add data
            if item.itemName == Name:
                item.changeSellPrice(Price)
                item.changeItemBarcode(Barcode)
                item.changeIemCost(Cost)
                item.changeinStock(Stock)
    return items


def init_OVVI(folder):
    wb = Workbook() 
    dest_filename = folder + 'Item-PLU with Data.xlsx'   
    ws1 = wb.active
    ws1.title = "Item-PLU" 
    ws2 = wb.create_sheet(title="ModifierGroups")   
    
    wb.save(filename = dest_filename)

def findItem(itemsList, itemName):
    index = 0
    for item in itemsList:
        if item.itemName == itemName:
            print(item.itemName)
            print(item.itemDepartment)
            print(index)
    index += 1