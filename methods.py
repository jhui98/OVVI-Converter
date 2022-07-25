# Jacob Hui - Clover Automation Functions

logo = """
   ____               _    _____                          _            
  / __ \             (_)  / ____|                        | |           
 | |  | |_   ____   ___  | |     ___  _ ____   _____ _ __| |_ ___ _ __ 
 | |  | \ \ / /\ \ / / | | |    / _ \| '_ \ \ / / _ \ '__| __/ _ \ '__|
 | |__| |\ V /  \ V /| | | |___| (_) | | | \ V /  __/ |  | ||  __/ |   
  \____/  \_/    \_/ |_|  \_____\___/|_| |_|\_/ \___|_|   \__\___|_|    -Jacob Hui                                                                      
"""

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def get_departments(clover_wb):
    ws = clover_wb["Categories"]
    departments = []
    for cell in ws[1]:
        departments.append(cell.value)
    departments.pop(0)
    # print(departments)
    return departments 

def item_department_dict(clover_wb, departments):
    item_dict = {}
    ws = clover_wb["Categories"] # select categories ws
    for col in range(len(departments)):
        dep = str(departments[col]) # save department name
        # print(dep)
        for cell in ws[get_column_letter(col+2)]: # for every cell within column index
            if (cell.value != None) and (cell.value != dep): # cell not empty and not department
                item = str(cell.value)
                item_dict[item] = {"department": dep}
    return item_dict

def initialItemIstance(clover_wb, departments):
    items = []
    ws = clover_wb["Categories"] # select categories ws
    for col in range(len(departments)):
        dep = str(departments[col]) # save department name
        for cell in ws[get_column_letter(col+2)]: # for every cell within column index
            if (cell.value != None) and (cell.value != dep): # cell not empty and not department
                item = str(cell.value)
                item = Ovvi(item, dep)
                items.append(item)

    ws = clover_wb["Items"] # select categories ws
    for row in ws.iter_rows(): # iter throguh rows
        rowIndex = str(row[0].row) # get a row number
        # get data from row based on col index
        Name = ws[f"B{rowIndex}"].value
        Price = ws[f"D{rowIndex}"].value
        Cost = ws[f"H{rowIndex}"].value
        Barcode = ws[f"J{rowIndex}"].value
        Stock = ws[f"L{rowIndex}"].value

        for item in items: # check items obj list for item name, add data
            if item.itemName == Name:
                item.changeSellPrice(Price)
                item.changeItemBarcode(Barcode)
                item.changeIemCost(Cost)
                item.changeinStock(Stock)
    return items


def init_OVVI(folder):
    wb = Workbook() 
    dest_filename = folder + 'Item-PLU with Data.xlsx'   
    ws1 = wb.active
    ws1.title = "Item-PLU" 
    ws2 = wb.create_sheet(title="ModifierGroups")   
    
    wb.save(filename = dest_filename)

def findItem(itemsList, itemName):
    index = 0
    for item in itemsList:
        if item.itemName == itemName:
            print(item.itemName)
            print(item.itemDepartment)
            print(index)
    index += 1


class Ovvi:
    def __init__(self, itemName, department):
        self.itemDepartment = department
        self.itemNumber = ""
        self.itemName = itemName
        self.modifierGroups = ""
        self.description = ""
        self.itemBarcode = ""
        self.itemCost = "0"
        self.itemSellPrice = "0"
        self.inStock = "0"
        self.tax1 = "TRUE"
        self.displayInMenu = "TRUE"
        self.isInventoryItem = "TRUE"
        self.isFoodStampable = "FALSE"
        self.beverageDeposit = ""

    # Item structure
        # item_name = {
        # department: "",
        # price: float(),
        # barcode: "",
        # cost: float(),
        # }
    
    # def change(self, ): # change 
    #     self. = 

    def changeitemDepartment(self, itemDepartment): # change itemDepartment
        self.itemDepartment = itemDepartment

    def changeitemName(self, itemName): # change itemName
        self.itemName = itemName

    def changeDescription(self, description): # change description
        self.description = description

    def changeItemBarcode(self, barcode): # change item barcode
        self.itemBarcode = barcode
    
    def changeIemCost(self, itemCost): # change item cost
        self.itemCost = itemCost

    def changeSellPrice(self, sellPrice): # change item sell price
        self.itemSellPrice = sellPrice

    def changeinStock(self, inStock): # change item inStock
        self.inStock = inStock





    
