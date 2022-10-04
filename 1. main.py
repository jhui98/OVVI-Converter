# Jacob Hui - Ovvi Converter

from queue import Empty
from clover_methods import get_departments, item_department_dict, initialItemIstance, Ovvi
from openpyxl import Workbook, load_workbook
from logo import logo
import time, os
from pathlib import Path
downloads_path = str(Path.home() / "Downloads")
# print(downloads_path)

run = "y"
while run != "n":
    print(logo)

    task = input("""
    A: Clover Conversion
    B: Barcode Leading Zero Fix
    C: Check My Work (Coming soon)

    Please choose an operation: """)

    inputFile = str(input("\nPlease drag and drop the file, then press Enter: "))
    
    # checking for formatting on input
    startIndex = inputFile.index(":") # starting index is char after drive indicator
    startIndex -= 1 # get one character back to get drive 
    if inputFile[-1] == "\'" or inputFile[-1] == "\"": # check if ends with quote
        endIndex = -1 # if quote at end, then end index one before
    else: # no quote no need to do anything
        endIndex = None
    inputFile = inputFile[startIndex:endIndex] # set inputFile to file path
    # print(inputFile) # sanity check
    print("\nProcessing, please wait...\n") # UX udpate
    
    if task == "A" or task == "a": # Clover to OVVI format change
        # load worksheet
        wb = load_workbook(inputFile)
        ws = wb.active
        # print(clover_wb.sheetnames) # sanity check

        departments = get_departments(wb) # get departments

        items_dict = item_department_dict(wb, departments) # create dictionary of items with assigned department
        items = initialItemIstance(wb, departments)

        # TODO: Look into this
        # from pprint import pprint
        # pprint(vars(your_object))
        
        # double checking
        # for item in items:
        #     if item.itemName == '':
        #         print(item.itemName)
        #         print(item.itemDepartment)
        #         print(item.itemSellPrice)
        #         print(item.itemBarcode)
        #         print(item.itemCost)
        #         print(item.itemStock)

        
        # store in new workbook
        wb = Workbook()
        ws = wb.active
        # create Item-PLU sheet
        ws.title = "Item-PLU"
        ws.append(["Department", "ItemNumber", "ItemName", "ModifierGroups", "Description", "Barcode", "Cost", "SellPrice", "InStock", "Tax1", "DisplayInMenu", "IsInventoryItem", "IsFoodStampItem", "BeveragesDeposit"])

        # add in clover items
        for item in items:
            ws.append([item.itemDepartment, item.itemNumber, item.itemName, item.modifierGroups, item.description, item.itemBarcode, item.itemCost, item.itemSellPrice, item.inStock, item.tax1, item.displayInMenu, item.isInventoryItem, item.isFoodStampable, item.beverageDeposit])
        
        # create ModifierGroups sheet
        wb.create_sheet("ModifierGroups")
        ws = wb["ModifierGroups"]
        ws.append(["Modifer Group Department", "Modifier Group Name", "Charged", "Modifier Department", "Modifier", "Price", "Min", "Max"])
    elif task == "B" or task == "b": # Barcode Leading Zero Fix
        # load worksheet
        wb = load_workbook(inputFile)
        ws = wb["Item-PLU"]
        
        for row in ws.iter_rows(): # iter throguh rows
            rowIndex = str(row[0].row) # get a row number
            # get data from row based on col index
            barcode = str(ws[f"F{rowIndex}"].value)
            char = "@"
            if barcode == "None": # do nothing if empty 
                pass 
            elif char in barcode:# do nothing since already has existing multi barcode
                pass 
            else: # if not empty and no @ in barcode
                barcode = f"{barcode}@0{barcode}"
                ws[f"F{rowIndex}"] = barcode
    elif task == "C" or task == "c": # Check My Work
        # TODO: length of dept
        # TODO: length of item name
        # TODO: length of description
        # TODO: special characters in barcode (not including @)
        # TODO: cost is number
        # TODO: sell price is number
        # TODO: tax1, display, inventory, foodstamp == TRUE or FALSE
        # TODO: beverage deposit either blank or numerical
        # get info from inspect elements
        

        pass

    # save
    timestr = time.strftime("%Y%m%d-%H%M%S")
    saveName = f"\Ovvi_Convert_Output_{timestr}.xlsx"
    wb.save(downloads_path + saveName)
    # print statement and ask for another operation
    print("Finished! Please check your downloads folder for the updated file.")
    print(f"Your new file is named: {saveName[1:]}\n")
    run = input("Would you like to process another operation? y or n: ")
    os.system("cls") # clear screen to reset after one operation